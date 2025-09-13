import pytest
from datetime import date

from hrms.db import SessionLocal
from hrms.models import Person
from hrms.contracts import add_contract, export_contracts_for_person
from openpyxl import load_workbook

def test_contracts_add_and_export(tmp_path):
    db = SessionLocal()
    try:
        p = db.query(Person).first()
        c = add_contract(db, p, "HĐ xác định thời hạn", date.today())
        out = tmp_path / f"contracts_{p.code}.xlsx"
        export_contracts_for_person(db, p, str(out))
        assert out.exists()
        assert out.stat().st_size > 0
        # Validate header and freeze panes
        wb = load_workbook(str(out))
        ws = wb.active
        assert [cell.value for cell in ws[1]] == ["Loại HĐ", "Từ ngày", "Đến ngày", "Ghi chú"]
        assert ws.freeze_panes == "A2"
    finally:
        db.close()
