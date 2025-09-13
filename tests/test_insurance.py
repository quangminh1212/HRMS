import pytest
from datetime import date

from hrms.db import SessionLocal
from hrms.models import Person
from hrms.insurance import add_insurance_event, export_insurance_to_excel
from openpyxl import load_workbook


def test_insurance_add_and_export(tmp_path):
    db = SessionLocal()
    try:
        p = db.query(Person).first()
        add_insurance_event(db, p, "Nghỉ thai sản", date.today())
        out = tmp_path / "bhxh.xlsx"
        export_insurance_to_excel(db, date(date.today().year,1,1), date(date.today().year,12,31), str(out))
        assert out.exists()
        assert out.stat().st_size > 0
        # Validate header and freeze panes
        wb = load_workbook(str(out))
        ws = wb.active
        assert [cell.value for cell in ws[1]] == ["Mã NV", "Họ tên", "Loại sự kiện", "Ngày", "Ghi chú"]
        assert ws.freeze_panes == "A2"
    finally:
        db.close()
