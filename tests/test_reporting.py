import pytest
from datetime import date

from hrms.db import SessionLocal
from hrms.reporting import compute_annual_summary, compute_demographics, export_report_to_excel
from openpyxl import load_workbook


def test_export_quick_report(tmp_path):
    db = SessionLocal()
    try:
        year = date.today().year
        summary = compute_annual_summary(db, year)
        demo = compute_demographics(db, date.today())
        out = tmp_path / f"bao_cao_nhanh_{year}.xlsx"
        export_report_to_excel(summary, demo, str(out))
        assert out.exists()
        assert out.stat().st_size > 0
        wb = load_workbook(str(out))
        assert 'Summary' in wb.sheetnames
        assert 'Demographics' in wb.sheetnames
        # Summary
        ws1 = wb['Summary']
        assert ws1['A1'].value == 'Year'
        assert ws1.freeze_panes == 'A2'
        # Demographics header
        ws2 = wb['Demographics']
        assert [c.value for c in ws2[1]] == ['Bucket', 'Count']
        assert ws2.freeze_panes == 'A2'
    finally:
        db.close()

import pytest
from datetime import date

from hrms.db import SessionLocal
from hrms.reporting import compute_annual_summary, compute_demographics


def test_reporting_summary_and_demo():
    db = SessionLocal()
    try:
        s = compute_annual_summary(db, date.today().year)
        assert set(["year","salary_increase","retirement","appointment","maternity","studying","leaving"]).issubset(s.keys())
        d = compute_demographics(db, date.today())
        for k in ["age_buckets","ethnicity","gender","llct","professional"]:
            assert k in d
    finally:
        db.close()
