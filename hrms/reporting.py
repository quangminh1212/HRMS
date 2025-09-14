from datetime import date
from collections import Counter
from typing import Dict, Any, List, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import func

from .models import Person, SalaryHistory, RetirementNotice, Appointment, EmailLog


def age_on(dob: date, as_of: date) -> int:
    if not dob:
        return -1
    years = as_of.year - dob.year - ((as_of.month, as_of.day) < (dob.month, dob.day))
    return years


def compute_annual_summary(db: Session, year: int) -> Dict[str, Any]:
    # Nâng lương: tính theo lịch sử lương trong năm
    inc = db.query(SalaryHistory).filter(
        SalaryHistory.effective_date >= date(year, 1, 1),
        SalaryHistory.effective_date <= date(year, 12, 31),
    ).count()

    # Nghỉ hưu: theo planned_date trong năm
    retire = db.query(RetirementNotice).filter(
        RetirementNotice.planned_date >= date(year, 1, 1),
        RetirementNotice.planned_date <= date(year, 12, 31),
    ).count()

    # Bổ nhiệm: theo start_date trong năm
    appoint = db.query(Appointment).filter(
        Appointment.start_date >= date(year, 1, 1),
        Appointment.start_date <= date(year, 12, 31),
    ).count()

    # Thai sản, đi học, thôi việc: dựa vào Person.status (đếm trạng thái hiện tại)
    maternity = db.query(Person).filter(Person.status.ilike('%thai sản%')).count()
    studying = db.query(Person).filter(Person.status.ilike('%đi học%')).count()
    leaving = db.query(Person).filter(Person.status.ilike('%thôi việc%')).count()

    return {
        "year": year,
        "salary_increase": inc,
        "retirement": retire,
        "appointment": appoint,
        "maternity": maternity,
        "studying": studying,
        "leaving": leaving,
    }


def compute_demographics(db: Session, as_of: date) -> Dict[str, Any]:
    people = db.query(Person).all()

    # Nhóm tuổi
    buckets = {"<30": 0, "30-40": 0, "41-50": 0, "51-60": 0, ">60": 0, "Unknown": 0}
    for p in people:
        a = age_on(p.dob, as_of)
        if a < 0:
            buckets["Unknown"] += 1
        elif a < 30:
            buckets["<30"] += 1
        elif a <= 40:
            buckets["30-40"] += 1
        elif a <= 50:
            buckets["41-50"] += 1
        elif a <= 60:
            buckets["51-60"] += 1
        else:
            buckets[">60"] += 1

    # Dân tộc, giới tính, LLCT, chuyên môn
    ethnicity = Counter([(p.ethnicity or "Khác").strip() for p in people])
    gender = Counter([(p.gender or "Khác").strip() for p in people])
    llct = Counter([(p.llct_level or "Chưa rõ").strip() for p in people])
    prof = Counter([(p.professional_level or "Chưa rõ").strip() for p in people])

    return {
        "age_buckets": buckets,
        "ethnicity": dict(ethnicity),
        "gender": dict(gender),
        "llct": dict(llct),
        "professional": dict(prof),
    }


def export_report_to_excel(summary: Dict[str, Any], demo: Dict[str, Any], path: str) -> None:
    from openpyxl import Workbook
    from .excel_utils import style_header, auto_filter_and_width

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Year", summary["year"]])
    ws1.append(["Salary Increase", summary["salary_increase"]])
    ws1.append(["Retirement", summary["retirement"]])
    ws1.append(["Appointment", summary["appointment"]])
    ws1.append(["Maternity (current)", summary["maternity"]])
    ws1.append(["Studying (current)", summary["studying"]])
    ws1.append(["Leaving (current)", summary["leaving"]])
    # Style + filter cho Summary
    style_header(ws1, header_row=1)
    auto_filter_and_width(ws1, header_row=1)

    ws2 = wb.create_sheet("Demographics")
    ws2.append(["Bucket", "Count"])
    for k, v in demo["age_buckets"].items():
        ws2.append([k, v])
    style_header(ws2, header_row=1)
    auto_filter_and_width(ws2, header_row=1)

    def write_counter(sheet_name: str, data: Dict[str, int]):
        ws = wb.create_sheet(sheet_name)
        ws.append(["Category", "Count"])
        for k, v in data.items():
            ws.append([k, v])
        style_header(ws, header_row=1)
        auto_filter_and_width(ws, header_row=1)

    write_counter("Ethnicity", demo["ethnicity"])
    write_counter("Gender", demo["gender"])
    write_counter("LLCT", demo["llct"])
    write_counter("Professional", demo["professional"])

    wb.save(path)


def compute_email_summary(db: Session, base_query) -> Dict[str, Any]:
    """Tính thống kê gọn cho EmailLog dựa trên base_query đã áp dụng filter.
    base_query là một sqlalchemy Query trên EmailLog.
    """
    q = base_query
    # status
    by_status = db.query(EmailLog.status, func.count(1)).filter(
        q._criterion if getattr(q, '_criterion', None) is not None else True
    ).group_by(EmailLog.status).all()
    # type
    by_type = db.query(EmailLog.type, func.count(1)).filter(
        q._criterion if getattr(q, '_criterion', None) is not None else True
    ).group_by(EmailLog.type).all()
    return {
        'by_status': [(k or '', v) for k, v in by_status],
        'by_type': [(k or '', v) for k, v in by_type],
    }


def export_retirement_alerts_to_excel(db: Session, persons_six: List[Person], persons_three: List[Person], path: str) -> None:
    from openpyxl import Workbook
    from .excel_utils import style_header, auto_filter_and_width, set_header_footer, set_date_format
    from .retirement import calculate_retirement_date
    from .settings_service import get_setting

    wb = Workbook()
    # Sheet for +6 months
    ws1 = wb.active
    ws1.title = "Retire+6m"
    ws1.append(["Mã NV", "Họ tên", "Đơn vị", "Chức vụ", "Ngày nghỉ hưu"])
    for p in persons_six:
        rd = calculate_retirement_date(p)
        ws1.append([p.code, p.full_name, p.unit.name if p.unit else "", p.position.name if p.position else "", rd])
    style_header(ws1, header_row=1)
    auto_filter_and_width(ws1, header_row=1)
    date_fmt = get_setting('XLSX_DATE_FORMAT', 'DD/MM/YYYY') or 'DD/MM/YYYY'
    set_date_format(ws1, date_columns=[5], start_row=2, fmt=date_fmt)

    # Sheet for +3 months
    ws2 = wb.create_sheet("Retire+3m")
    ws2.append(["Mã NV", "Họ tên", "Đơn vị", "Chức vụ", "Ngày nghỉ hưu"])
    for p in persons_three:
        rd = calculate_retirement_date(p)
        ws2.append([p.code, p.full_name, p.unit.name if p.unit else "", p.position.name if p.position else "", rd])
    style_header(ws2, header_row=1)
    auto_filter_and_width(ws2, header_row=1)
    set_date_format(ws2, date_columns=[5], start_row=2, fmt=date_fmt)

    org = get_setting('ORG_NAME', None)
    set_header_footer(ws1, title='Danh sách nghỉ hưu (+6 tháng)', org=org)
    set_header_footer(ws2, title='Danh sách nghỉ hưu (+3 tháng)', org=org)

    wb.save(path)
