from datetime import date, timedelta
from math import floor
from typing import List, Optional, Dict, Any

from sqlalchemy.orm import Session
from sqlalchemy import func

from .models import Person, SalaryRank, SalaryStep, SalaryHistory, Allowance, Unit, Position


def months_between(d1: date, d2: date) -> int:
    """Số tháng tròn giữa d1 -> d2 (floor)."""
    if d2 < d1:
        return -months_between(d2, d1)
    return (d2.year - d1.year) * 12 + (d2.month - d1.month) - (0 if d2.day >= d1.day else 1)


def latest_salary_history(db: Session, person_id: int) -> Optional[SalaryHistory]:
    return (
        db.query(SalaryHistory)
        .filter(SalaryHistory.person_id == person_id)
        .order_by(SalaryHistory.effective_date.desc())
        .first()
    )


def max_step_for_rank(db: Session, rank_id: int) -> Optional[int]:
    return db.query(func.max(SalaryStep.step)).filter(SalaryStep.rank_id == rank_id).scalar()


def get_step_info(db: Session, rank_id: int, step: int) -> Optional[SalaryStep]:
    return (
        db.query(SalaryStep)
        .filter(SalaryStep.rank_id == rank_id, SalaryStep.step == step)
        .first()
    )


def compute_next_for_person(db: Session, person: Person, as_of: date) -> Optional[Dict[str, Any]]:
    """
    Trả về dict thông tin đề xuất nâng lương tiếp theo hoặc None nếu chưa đến hạn.
    Quy tắc:
    - Nếu chưa ở bậc cuối: đủ số tháng tối thiểu (36/24 tuỳ ngạch) -> tăng 1 bậc.
    - Nếu ở bậc cuối: đủ 36 tháng (chuyên viên & tương đương) hoặc 24 tháng (nhân viên, thủ quỹ)
      -> phụ cấp thâm niên vượt khung 5%, sau đó mỗi 12 tháng +1%.
    - Nếu SalaryHistory.note chứa từ khoá trì hoãn/kỷ luật -> bỏ qua.
    """
    hist = latest_salary_history(db, person.id)
    if not hist:
        return None

    if hist.note:
        low = hist.note.lower()
        if any(k in low for k in ["ky luat", "kỷ luật", "delay", "trì hoãn", "keo dai", "kéo dài"]):
            return None

    step_info = get_step_info(db, hist.rank_id, hist.step)
    if not step_info:
        return None
    months = months_between(hist.effective_date, as_of)

    max_step = max_step_for_rank(db, hist.rank_id) or hist.step
    rank = db.get(SalaryRank, hist.rank_id)
    rank_level = (rank.level or "").lower() if rank else ""

    # Ngưỡng tháng theo ngạch
    threshold = 36
    if "nhan vien" in rank_level or "thu quy" in rank_level or "nhân viên" in rank_level or "thủ quỹ" in rank_level:
        threshold = 24

    if hist.step < max_step:
        # chưa ở bậc cuối
        if months >= step_info.min_months:
            next_step = hist.step + 1
            next_step_info = get_step_info(db, hist.rank_id, next_step)
            if not next_step_info:
                return None
            due_months = step_info.min_months
            due_date = add_months(hist.effective_date, due_months)
            return {
                "type": "step",
                "person": person,
                "current_step": hist.step,
                "current_coef": hist.coefficient,
                "next_step": next_step,
                "next_coef": next_step_info.coefficient,
                "last_effective": hist.effective_date,
                "due_date": due_date,
                "months_waited": months,
            }
        return None
    else:
        # bậc cuối -> thâm niên vượt khung
        if months >= threshold:
            extra_years = max(0, floor((months - threshold) / 12))
            percent = 5 + extra_years
            due_date = add_months(hist.effective_date, threshold)
            return {
                "type": "over_limit",
                "person": person,
                "current_step": hist.step,
                "current_coef": hist.coefficient,
                "allowance_percent": percent,
                "last_effective": hist.effective_date,
                "due_date": due_date,
                "months_waited": months,
            }
        return None


def add_months(d: date, m: int) -> date:
    y = d.year + (d.month - 1 + m) // 12
    mo = (d.month - 1 + m) % 12 + 1
    # ngày cuối tháng nếu ngày cũ vượt quá số ngày trong tháng mới
    from calendar import monthrange

    last_day = monthrange(y, mo)[1]
    day = min(d.day, last_day)
    return date(y, mo, day)


def list_due_in_window(
    db: Session,
    start_date: date,
    end_date: date,
    unit_id: Optional[int] = None,
    position_id: Optional[int] = None,
) -> List[Dict[str, Any]]:
    q = db.query(Person)
    if unit_id:
        q = q.filter(Person.unit_id == unit_id)
    if position_id:
        q = q.filter(Person.position_id == position_id)
    people = q.all()

    results: List[Dict[str, Any]] = []
    for p in people:
        # Tính mốc due tại as_of = end_date để đảm bảo bao phủ cửa sổ
        info = compute_next_for_person(db, p, end_date)
        if not info:
            continue
        due = info["due_date"]
        if start_date <= due <= end_date:
            # enrich basic fields
            info["full_name"] = p.full_name
            info["unit_name"] = p.unit.name if p.unit else ""
            info["position_name"] = p.position.name if p.position else ""
            info["days_left"] = (due - start_date).days
            results.append(info)
    # sắp xếp theo ngày đến hạn
    results.sort(key=lambda x: x["due_date"]) 
    return results


def export_due_to_excel(items: List[Dict[str, Any]], file_path: str) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Nang luong"
    headers = [
        "Họ tên", "Đơn vị", "Chức vụ", "Loại", "Bậc hiện tại", "Hệ số hiện tại",
        "Bậc dự kiến", "Hệ số dự kiến", "% vượt khung", "Ngày hiệu lực gần nhất", "Ngày dự kiến",
    ]
    ws.append(headers)
    for it in items:
        ws.append([
            it.get("full_name", ""),
            it.get("unit_name", ""),
            it.get("position_name", ""),
            "Tăng bậc" if it.get("type") == "step" else "Vượt khung",
            it.get("current_step", ""),
            it.get("current_coef", ""),
            it.get("next_step", "") if it.get("type") == "step" else "",
            it.get("next_coef", "") if it.get("type") == "step" else "",
            it.get("allowance_percent", "") if it.get("type") == "over_limit" else "",
            it.get("last_effective", ""),
            it.get("due_date", ""),
        ])
    wb.save(file_path)


def export_salary_history_for_person(db: Session, person: Person, file_path: str) -> None:
    """Xuất toàn bộ lịch sử lương của một nhân sự ra Excel, có định dạng và biểu đồ hệ số theo thời gian."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary history"
    headers = ["Mã NV", "Họ tên", "Ngày hiệu lực", "Ngạch", "Bậc", "Hệ số", "Ghi chú"]
    ws.append(headers)

    # Header style + freeze panes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    hists = (
        db.query(SalaryHistory)
        .filter(SalaryHistory.person_id == person.id)
        .order_by(SalaryHistory.effective_date.asc())
        .all()
    )

    # Tránh N+1: cache rank code theo id nếu cần
    rank_code_cache: Dict[int, str] = {}

    for h in hists:
        rid = h.rank_id
        if rid not in rank_code_cache:
            r = db.get(SalaryRank, rid)
            rank_code_cache[rid] = (r.code if r else "")
        ws.append([
            person.code or "",
            person.full_name or "",
            h.effective_date,
            rank_code_cache.get(rid, ""),
            h.step,
            h.coefficient,
            h.note or "",
        ])

    # Auto filter toàn bảng
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    # Auto width cột
    widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx, val in enumerate(row):
            l = len(str(val)) if val is not None else 0
            if l > widths[idx]:
                widths[idx] = l
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 40)

    # Biểu đồ hệ số theo thời gian nếu có dữ liệu
    if ws.max_row > 1:
        chart = LineChart()
        chart.title = f"Biểu đồ hệ số theo thời gian - {person.full_name or ''}"
        chart.y_axis.title = "Hệ số"
        chart.x_axis.title = "Ngày hiệu lực"
        data = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)  # cột Hệ số
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # cột Ngày hiệu lực
        chart.set_categories(cats)
        ws.add_chart(chart, "I2")

    wb.save(file_path)


def export_salary_histories_for_people(db: Session, people: List[Person], file_path: str) -> None:
    """Xuất lịch sử lương cho danh sách nhân sự (một sheet tổng hợp) với định dạng."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary histories"
    headers = ["Mã NV", "Họ tên", "Ngày hiệu lực", "Ngạch", "Bậc", "Hệ số", "Ghi chú"]
    ws.append(headers)

    # Header style + freeze panes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    if not people:
        ws.auto_filter.ref = "A1:G1"
        wb.save(file_path)
        return

    ids = [p.id for p in people]
    person_map = {p.id: p for p in people}

    # Lấy toàn bộ lịch sử cho các nhân sự đã chọn
    hists = (
        db.query(SalaryHistory)
        .filter(SalaryHistory.person_id.in_(ids))
        .order_by(SalaryHistory.person_id.asc(), SalaryHistory.effective_date.asc())
        .all()
    )

    # Cache mã ngạch để tránh N+1
    rank_ids = sorted({h.rank_id for h in hists})
    code_map: Dict[int, str] = {}
    if rank_ids:
        for r in db.query(SalaryRank).filter(SalaryRank.id.in_(rank_ids)).all():
            code_map[r.id] = r.code or ""

    for h in hists:
        p = person_map.get(h.person_id)
        if not p:
            continue
        ws.append([
            p.code or "",
            p.full_name or "",
            h.effective_date,
            code_map.get(h.rank_id, ""),
            h.step,
            h.coefficient,
            h.note or "",
        ])

    # Auto filter + auto width
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx, val in enumerate(row):
            l = len(str(val)) if val is not None else 0
            if l > widths[idx]:
                widths[idx] = l
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 40)

    wb.save(file_path)
