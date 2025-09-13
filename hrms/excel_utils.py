from __future__ import annotations
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def prepare_workbook_with_template(template_name: Optional[str], title: str, headers: List[str]) -> Tuple[Workbook, any]:
    """
    Tạo Workbook/Worksheet từ template nếu có, ngược lại tạo mới với tiêu đề và header.
    - Nếu template tồn tại nhưng không có header khớp, sẽ ghi header vào dòng 1.
    """
    if template_name:
        tpl_path = Path('templates') / 'xlsx' / template_name
        if tpl_path.exists():
            wb = load_workbook(str(tpl_path))
            ws = wb.active
            # Đảm bảo header
            first_row = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
            if first_row != headers:
                for idx, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=idx, value=h)
            return wb, ws
    # Fallback: tạo mới
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    return wb, ws


def style_header(ws, header_row: int = 1, fill_color: str = '4F81BD', font_color: str = 'FFFFFF', bold: bool = True,
                 freeze: bool = True) -> None:
    header_font = Font(bold=bold, color=font_color)
    header_fill = PatternFill('solid', fgColor=fill_color)
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if freeze:
        ws.freeze_panes = f"A{header_row + 1}"


def auto_filter_and_width(ws, header_row: int = 1, min_width: int = 10, max_width: int = 40) -> None:
    # Auto filter toàn bảng
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    # Tính toán độ rộng cột
    widths = [len(str(ws.cell(row=header_row, column=i + 1).value or "")) for i in range(ws.max_column)]
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        for idx, val in enumerate(row):
            l = len(str(val)) if val is not None else 0
            if l > widths[idx]:
                widths[idx] = l
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, min_width), max_width)


def set_date_format(ws, date_columns: List[int], start_row: int = 2, fmt: str = 'DD/MM/YYYY') -> None:
    for col in date_columns:
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = fmt


def set_number_format(ws, number_columns: List[int], start_row: int = 2, fmt: str = '0.00') -> None:
    for col in number_columns:
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = fmt
